import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

from .analyzer import make_unique_columns

FOOTER_KEYWORDS = {
    "합계",
    "총계",
    "소계",
    "평균",
    "COUNT",
    "TOTAL",
    "건수",
}

def load_workbook_file(path):
    """
    Excel Workbook 로드
    """
    return load_workbook(path)

def worksheet_to_dataframe(ws, analyzer):
    """
    Worksheet -> DataFrame

    합계/소계 등 요약 행은 표 중간에 나올 수 있으므로 데이터 읽기를
    중단시키지 않는다. 대신 footer_rows 에 위치를 기록해 두고,
    마스킹 단계에서만 건너뛴다 (행 위치는 워크시트와 그대로 정렬 유지).
    """
    rows = list(ws.values)

    if not rows:
        return pd.DataFrame(), 1, set()

    header_row = analyzer.find_header_row(ws)

    columns = rows[header_row - 1]

    data = []
    footer_rows = set()

    for row in rows[header_row:]:

        if is_blank_row(row):
            break

        if is_footer_row(row):
            footer_rows.add(len(data))

        data.append(row)

    # 중복 헤더가 있으면 pandas 접근이 깨지므로 유일한 라벨로 만든다.
    # (다시 엑셀에 쓸 때는 열 위치 기준이라 원본 헤더는 그대로 유지된다.)
    df = pd.DataFrame(data, columns=make_unique_columns(list(columns)))

    return df, header_row, footer_rows

def dataframe_to_worksheet(ws, df, header_row):
    """
    DataFrame 값을 Worksheet에 다시 기록
    서식(폰트, 색상, 테두리, 병합 등)은 유지하고
    셀 값만 변경합니다.
    """

    for row_idx, row in enumerate(
        df.itertuples(index=False),
        start=header_row + 1
    ):

        for col_idx, value in enumerate(row, start=1):

            cell = ws.cell(row=row_idx, column=col_idx)

            # 수식이면 건드리지 않는다.
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue

            # pd.NA / NaN / NaT 등은 openpyxl이 쓸 수 없으므로 빈 셀로 변환
            if pd.isna(value):
                value = None

            cell.value = value

def save_workbook(workbook, input_path, suffix):
    input_path = Path(input_path)

    output_path = input_path.with_name(
        f"{input_path.stem}{suffix}{input_path.suffix}"
    )

    workbook.save(output_path)

    return str(output_path)

def is_blank_row(row):
    """
    완전히 빈 행인지 여부 (표의 실제 끝을 의미)
    """
    return all(
        value is None or str(value).strip() == ""
        for value in row
    )


def is_footer_row(row):
    """
    합계/소계 등 요약 행인지 여부.

    표 끝을 의미하지는 않는다 (진료과별 소계처럼 표 중간에도 나올 수 있음).
    """
    first = row[0]

    if isinstance(first, str):
        first = first.strip().upper()

        if any(
            first.startswith(keyword)
            for keyword in FOOTER_KEYWORDS
        ):
            return True

    return False