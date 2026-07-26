# -*- coding: utf-8 -*-
import os
import io
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelSetProcessor:
    """
    엑셀/CSV 데이터의 두 컬럼 간 집합 연산(교집합, 합집합, 차집합 A-B, 차집합 B-A, 대칭차집합)을
    수행하고 스타일링된 엑셀 보고서를 생성하는 코어 엔진
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.filename = os.path.basename(file_path)
        self.ext = os.path.splitext(file_path)[1].lower()
        self._excel_file = None
        
        if self.ext in ['.xlsx', '.xls']:
            self._excel_file = pd.ExcelFile(file_path, engine='openpyxl' if self.ext == '.xlsx' else None)

    def get_sheet_names(self) -> List[str]:
        """시트 목록 반환"""
        if self.ext == '.csv':
            return ["CSV_Data"]
        elif self._excel_file:
            return self._excel_file.sheet_names
        return []

    def get_columns(self, sheet_name: str) -> List[str]:
        """해당 시트의 컬럼 목록 반환"""
        df = self.load_dataframe(sheet_name, nrows=5)
        return list(df.columns)

    def load_dataframe(self, sheet_name: str, nrows: Optional[int] = None) -> pd.DataFrame:
        """데이터프레임 로드"""
        if self.ext == '.csv':
            try:
                df = pd.read_csv(self.file_path, nrows=nrows, encoding='utf-8-sig')
            except Exception:
                df = pd.read_csv(self.file_path, nrows=nrows, encoding='cp949')
        else:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=nrows, engine='openpyxl' if self.ext == '.xlsx' else None)
        
        df.columns = [str(col).strip() for col in df.columns]
        return df

    def analyze_sets(
        self,
        sheet_a: str,
        col_a: str,
        sheet_b: Optional[str] = None,
        col_b: Optional[str] = None,
        case_sensitive: bool = False,
        trim_space: bool = True,
        drop_empty: bool = True,
        include_extra_cols: bool = True
    ) -> Dict[str, Any]:
        """두 컬럼 집합 비교 연산 수행"""
        if not sheet_b:
            sheet_b = sheet_a
        if not col_b:
            col_b = col_a

        df_a = self.load_dataframe(sheet_a)
        df_b = self.load_dataframe(sheet_b) if (sheet_b != sheet_a or self.ext == '.csv') else df_a

        if col_a not in df_a.columns:
            raise ValueError(f"시트 '{sheet_a}'에 컬럼 '{col_a}'가 존재하지 않습니다.")
        if col_b not in df_b.columns:
            raise ValueError(f"시트 '{sheet_b}'에 컬럼 '{col_b}'가 존재하지 않습니다.")

        def normalize_val(val: Any) -> str:
            if pd.isna(val) or val is None:
                return ""
            s = str(val)
            if trim_space:
                s = s.strip()
            if not case_sensitive:
                s = s.lower()
            return s

        series_a_raw = df_a[col_a]
        series_b_raw = df_b[col_b]

        items_a = []
        for idx, val in series_a_raw.items():
            norm = normalize_val(val)
            if drop_empty and norm == "":
                continue
            items_a.append({
                'norm': norm,
                'raw': str(val).strip() if trim_space and pd.notna(val) else str(val),
                'row_idx': idx + 2,
                'row_data': df_a.iloc[idx].to_dict() if include_extra_cols else {}
            })

        items_b = []
        for idx, val in series_b_raw.items():
            norm = normalize_val(val)
            if drop_empty and norm == "":
                continue
            items_b.append({
                'norm': norm,
                'raw': str(val).strip() if trim_space and pd.notna(val) else str(val),
                'row_idx': idx + 2,
                'row_data': df_b.iloc[idx].to_dict() if include_extra_cols else {}
            })

        set_norm_a = set(item['norm'] for item in items_a)
        set_norm_b = set(item['norm'] for item in items_b)

        norm_to_raw_a = {item['norm']: item['raw'] for item in items_a}
        norm_to_raw_b = {item['norm']: item['raw'] for item in items_b}

        # 1. 교집합 (A ∩ B)
        norm_intersection = set_norm_a.intersection(set_norm_b)
        # 2. A 전용 / 차집합 A (A - B)
        norm_a_only = set_norm_a.difference(set_norm_b)
        # 3. B 전용 / 차집합 B (B - A)
        norm_b_only = set_norm_b.difference(set_norm_a)
        # 4. 통합 대칭차집합 (A Δ B)
        norm_sym_diff = norm_a_only.union(norm_b_only)
        # 5. 합집합 (A ∪ B)
        norm_union = set_norm_a.union(set_norm_b)

        def build_result_list(norm_set: set, origin_filter: Optional[str] = None) -> List[Dict[str, Any]]:
            results = []
            sorted_norms = sorted(list(norm_set))
            for norm in sorted_norms:
                in_a = norm in set_norm_a
                in_b = norm in set_norm_b
                
                raw_val = norm_to_raw_a.get(norm, norm_to_raw_b.get(norm, norm))
                
                if origin_filter == 'A_ONLY' and not (in_a and not in_b):
                    continue
                if origin_filter == 'B_ONLY' and not (in_b and not in_a):
                    continue

                if in_a and not in_b:
                    origin = "A전용(차집합A)"
                elif in_b and not in_a:
                    origin = "B전용(차집합B)"
                else:
                    origin = "공통(교집합)"

                results.append({
                    'val': raw_val,
                    'origin': origin,
                    'in_a': "O" if in_a else "X",
                    'in_b': "O" if in_b else "X",
                })
            return results

        list_intersection = build_result_list(norm_intersection)
        list_a_only = build_result_list(norm_a_only, origin_filter='A_ONLY')
        list_b_only = build_result_list(norm_b_only, origin_filter='B_ONLY')
        list_sym_diff = list_a_only + list_b_only
        list_union = build_result_list(norm_union)

        stats = {
            'col_a_name': f"{sheet_a} -> {col_a}",
            'col_b_name': f"{sheet_b} -> {col_b}",
            'total_a_rows': len(df_a[col_a]),
            'total_b_rows': len(df_b[col_b]),
            'unique_a_count': len(set_norm_a),
            'unique_b_count': len(set_norm_b),
            'intersection_count': len(norm_intersection),
            'a_only_count': len(norm_a_only),
            'b_only_count': len(norm_b_only),
            'sym_diff_count': len(norm_sym_diff),
            'union_count': len(norm_union),
            'intersection_pct_a': round(len(norm_intersection) / max(len(set_norm_a), 1) * 100, 1),
            'intersection_pct_b': round(len(norm_intersection) / max(len(set_norm_b), 1) * 100, 1),
            'case_sensitive': case_sensitive,
            'trim_space': trim_space,
            'drop_empty': drop_empty
        }

        return {
            'stats': stats,
            'intersection': list_intersection,
            'a_only': list_a_only,
            'b_only': list_b_only,
            'sym_diff': list_sym_diff,
            'union': list_union
        }

    def export_excel_report(self, analysis_result: Dict[str, Any], output_path: str):
        """서식이 적용된 멀티시트 Excel 파일 내보내기"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        stats = analysis_result['stats']

        COLOR_PRIMARY = "1E293B"     
        COLOR_BLUE_HEADER = "2563EB"
        COLOR_YELLOW_HEADER = "D97706"
        COLOR_RED_HEADER = "DC2626"
        COLOR_PURPLE_HEADER = "7C3AED"
        COLOR_GREEN_HEADER = "059669"

        font_title = Font(name="맑은 고딕", size=16, bold=True, color="1E293B")
        font_subtitle = Font(name="맑은 고딕", size=11, bold=True, color="475569")
        font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="맑은 고딕", size=10, bold=True)
        font_regular = Font(name="맑은 고딕", size=10)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin_border_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # -------------------------------------------------------------
        # 시트 1: 요약_대시보드
        # -------------------------------------------------------------
        ws_summary = wb.create_sheet(title="요약_대시보드")
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = "📊 엑셀 컬럼 집합 분석 요약 보고서"
        ws_summary["A1"].font = font_title
        ws_summary.row_dimensions[1].height = 30

        ws_summary["A2"] = f"분석 대상: [A컬럼] {stats['col_a_name']}  vs  [B컬럼] {stats['col_b_name']}"
        ws_summary["A2"].font = font_subtitle
        ws_summary.row_dimensions[2].height = 20

        summary_headers = ["연산 구분", "항목 설명", "고유 데이터 수", "비율 (A 기준)", "비율 (B 기준)"]
        ws_summary.append([])
        ws_summary.append(summary_headers)
        ws_summary.row_dimensions[4].height = 26

        for col_num in range(1, 6):
            cell = ws_summary.cell(row=4, column=col_num)
            cell.font = font_header
            cell.fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
            cell.alignment = align_center
            cell.border = thin_border

        summary_rows = [
            ("컬럼 A 전체", f"{stats['col_a_name']} 고유 항목 수", stats['unique_a_count'], "100.0%", f"{round(stats['unique_a_count']/max(stats['unique_b_count'],1)*100, 1)}%"),
            ("컬럼 B 전체", f"{stats['col_b_name']} 고유 항목 수", stats['unique_b_count'], f"{round(stats['unique_b_count']/max(stats['unique_a_count'],1)*100, 1)}%", "100.0%"),
            ("🔵 교집합 (A ∩ B)", "A와 B 양쪽에 모두 존재", stats['intersection_count'], f"{stats['intersection_pct_a']}%", f"{stats['intersection_pct_b']}%"),
            ("🟡 A전용 (차집합 A)", "A에만 있고 B에는 없음", stats['a_only_count'], f"{round(stats['a_only_count']/max(stats['unique_a_count'],1)*100, 1)}%", "-"),
            ("🔴 B전용 (차집합 B)", "B에만 있고 A에는 없음", stats['b_only_count'], "-", f"{round(stats['b_only_count']/max(stats['unique_b_count'],1)*100, 1)}%"),
            ("🟣 통합 대칭차집합", "불일치 항목 전체 (A전용 + B전용)", stats['sym_diff_count'], f"{round(stats['sym_diff_count']/max(stats['unique_a_count'],1)*100, 1)}%", f"{round(stats['sym_diff_count']/max(stats['unique_b_count'],1)*100, 1)}%"),
            ("🟢 합집합 (A ∪ B)", "A 또는 B에 존재하는 전체 항목", stats['union_count'], "-", "-")
        ]

        row_idx = 5
        for row_data in summary_rows:
            ws_summary.append(list(row_data))
            ws_summary.row_dimensions[row_idx].height = 22
            for c_idx in range(1, 6):
                cell = ws_summary.cell(row=row_idx, column=c_idx)
                cell.font = font_bold if c_idx == 1 else font_regular
                cell.border = thin_border
                if c_idx in [1, 3, 4, 5]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
            row_idx += 1

        def create_data_sheet(sheet_title: str, header_color: str, data_list: List[Dict[str, Any]]):
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            headers = ["번호", "데이터 값", "구분 (출처)", "A컬럼 존재", "B컬럼 존재"]
            ws.append(headers)
            ws.row_dimensions[1].height = 26

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                cell.alignment = align_center
                cell.border = thin_border

            for idx, item in enumerate(data_list, 1):
                row = [idx, item['val'], item['origin'], item['in_a'], item['in_b']]
                ws.append(row)
                curr_row = idx + 1
                ws.row_dimensions[curr_row].height = 20
                
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=curr_row, column=c_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    if c_idx in [1, 3, 4, 5]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len * 1.4 + 4, 14)

        create_data_sheet("교집합", COLOR_BLUE_HEADER, analysis_result['intersection'])
        create_data_sheet("A전용(차집합A)", COLOR_YELLOW_HEADER, analysis_result['a_only'])
        create_data_sheet("B전용(차집합B)", COLOR_RED_HEADER, analysis_result['b_only'])
        create_data_sheet("통합_대칭차집합", COLOR_PURPLE_HEADER, analysis_result['sym_diff'])
        create_data_sheet("합집합", COLOR_GREEN_HEADER, analysis_result['union'])

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len * 1.5 + 4, 16)

        wb.save(output_path)
        return output_path
