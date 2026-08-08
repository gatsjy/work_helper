# -*- coding: utf-8 -*-
"""
비식별화 어댑터 테스트.

vendored hkdeid 패키지의 마스킹 로직 자체는 원본 저장소 테스트가 담당한다.
여기서는 work_helper 통합 계층 — 키 안정성, 결과 리포트, 안전장치 — 를 본다.
"""
import openpyxl
import pytest

import deid_service as ds
from hkdeid.config import HKDeIDConfig


HEADERS = ["성명", "등록번호", "전화번호", "주민등록번호", "입원일", "비고"]
ROWS = [
    ["홍길동", "100001", "010-1234-5678", "900101-1234567", "2026-01-05", "보호자 010-9999-8888"],
    ["김철수", "100002", "010-2222-3333", "880202-2345678", "2026-01-07", "특이사항 없음"],
    ["홍길동", "100001", "010-1234-5678", "900101-1234567", "2026-02-01", ""],
]


def make_workbook(path, headers=HEADERS, rows=ROWS, title_rows=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "환자목록"
    if title_rows:
        ws.append(["HK병원 입원 현황"])
        ws.append(["조회기간: 2026-01-01 ~ 2026-02-28"])
        ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


@pytest.fixture
def source(tmp_path):
    return str(make_workbook(tmp_path / "patients.xlsx"))


def data_rows(path):
    ws = openpyxl.load_workbook(path).worksheets[0]
    return [r for r in ws.iter_rows(min_row=5, values_only=True)]


class TestKeyHandling:

    def test_key_path_is_absolute_and_stable(self):
        first = ds.resolve_key_file()
        assert first.is_absolute()
        assert first == ds.resolve_key_file()

    def test_masker_uses_the_patched_cached_key(self):
        import hkdeid.masker as masker
        assert masker.get_secret_key is ds._cached_get_secret_key

    def test_key_is_read_once_not_per_cell(self, source, tmp_path, monkeypatch):
        # 회귀: 원본은 셀 하나 가명화할 때마다 키 파일을 다시 열었다.
        ds._cached_get_secret_key()          # 캐시 예열
        calls = []
        original = ds._original_get_secret_key
        monkeypatch.setattr(
            ds, "_original_get_secret_key",
            lambda: (calls.append(1), original())[1]
        )
        ds.deidentify_workbook(source, str(tmp_path / "out.xlsx"))
        assert calls == []


class TestPreview:

    def test_detects_columns_without_writing(self, source, tmp_path):
        before = set(p.name for p in tmp_path.iterdir())
        report = ds.analyze_workbook(source)
        assert set(p.name for p in tmp_path.iterdir()) == before

        sheet = report.sheets[0]
        assert sheet.title == "환자목록"
        assert sheet.detected["patient_name"] == ["성명"]
        assert sheet.detected["patient_id"] == ["등록번호"]
        assert sheet.detected["phone"] == ["전화번호"]
        assert sheet.pii_columns >= 4

    def test_skips_title_rows_and_finds_header(self, source):
        report = ds.analyze_workbook(source)
        assert report.sheets[0].header_row == 4
        assert report.sheets[0].rows == len(ROWS)


class TestDeidentify:

    def test_masks_pii_and_preserves_structure(self, source, tmp_path):
        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out)

        ws = openpyxl.load_workbook(out).worksheets[0]
        assert ws.title == "환자목록"
        assert ws.cell(row=4, column=1).value == "성명"      # 헤더 보존
        assert ws.cell(row=1, column=1).value == "HK병원 입원 현황"

        rows = data_rows(out)
        assert rows[0][0].startswith("PATIENT_")
        assert rows[0][1].startswith("MRN_")
        assert rows[0][2] == "010-****-5678"
        assert rows[0][3] == "900101-*******"
        assert "홍길동" not in str(rows)
        assert "100001" not in str(rows)

    def test_same_patient_gets_same_pseudonym(self, source, tmp_path):
        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out)
        rows = data_rows(out)
        # 1행과 3행은 같은 환자다
        assert rows[0][0] == rows[2][0]
        assert rows[0][1] == rows[2][1]

    def test_deterministic_across_runs(self, source, tmp_path):
        a = str(tmp_path / "a.xlsx")
        b = str(tmp_path / "b.xlsx")
        ds.deidentify_workbook(source, a)
        ds.deidentify_workbook(source, b)
        assert data_rows(a) == data_rows(b)

    def test_free_text_pii_is_masked(self, source, tmp_path):
        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out)
        assert "010-9999-8888" not in str(data_rows(out))

    def test_date_shift_preserves_intervals(self, source, tmp_path):
        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out)
        rows = data_rows(out)
        # 원본 2026-01-05 와 2026-02-01 은 27일 차이 — 이동 후에도 같아야 한다
        assert (rows[2][4] - rows[0][4]).days == 27

    def test_reports_unique_patients(self, source, tmp_path):
        report = ds.deidentify_workbook(source, str(tmp_path / "out.xlsx"))
        assert report.unique_patients == 2      # 3행이지만 환자는 2명
        assert report.total_rows == 3
        assert report.processed_sheets == 1


class TestSafetyGuards:

    def test_refuses_to_overwrite_source(self, source):
        with pytest.raises(ValueError, match="원본"):
            ds.deidentify_workbook(source, source)

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ds.analyze_workbook(str(tmp_path / "nope.xlsx"))

    def test_warns_loudly_when_no_pii_detected(self, tmp_path):
        # 비식별화 툴에서 가장 위험한 실패: 아무것도 못 가리고 "성공"하는 것.
        path = make_workbook(
            tmp_path / "plain.xlsx",
            headers=["항목", "수량", "금액"],
            rows=[["볼펜", 10, 5000]],
            title_rows=False,
        )
        report = ds.analyze_workbook(str(path))
        assert report.total_pii_columns == 0
        assert any("개인정보 열을 하나도 찾지 못했" in w for w in report.warnings)

    def test_cancellation_writes_no_output(self, source, tmp_path):
        out = tmp_path / "cancelled.xlsx"
        with pytest.raises(ds.DeIdCancelled):
            ds.deidentify_workbook(
                source, str(out), should_cancel=lambda: True
            )
        assert not out.exists()

    def test_progress_reaches_completion(self, source, tmp_path):
        seen = []
        ds.deidentify_workbook(
            source, str(tmp_path / "out.xlsx"),
            progress=lambda p, m: seen.append(p),
        )
        assert seen and seen[-1] == 100
        assert seen == sorted(seen), "진행률이 뒤로 갔다"


class TestConfig:

    def test_disabled_masker_leaves_values_alone(self, source, tmp_path):
        config = HKDeIDConfig.default()
        config.mask_phone = False

        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out, config=config)
        assert data_rows(out)[0][2] == "010-1234-5678"

    def test_custom_date_shift(self, source, tmp_path):
        config = HKDeIDConfig.default()
        config.date_shift_days = -10

        out = str(tmp_path / "out.xlsx")
        ds.deidentify_workbook(source, out, config=config)
        assert data_rows(out)[0][4].strftime("%Y-%m-%d") == "2025-12-26"
